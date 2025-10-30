"""
qualcoder_core.py (updated)
Core functions for the three-stage qualitative coding pipeline with
an offline NLP keyword suggestion helper (TF-IDF based).
"""

from pathlib import Path
from typing import List, Dict, Tuple, Optional
import re
import json
import logging
import datetime
import pandas as pd
import PyPDF2
import docx
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# New: TF-IDF
from sklearn.feature_extraction.text import TfidfVectorizer

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


DEFAULT_CODEBOOK = {
    "Professional identity and experience": ["professor", "lecturer", "teacher", "years of experience", "experience"],
    "Use of Learning Management Systems": ["lms", "learning management", "blackboard", "moodle"],
    "Use of multimedia resources": ["youtube", "video", "multimedia", "visual"],
    "Use of social media for education": ["whatsapp", "social media", "facebook", "telegram"],
    "Virtual teaching platforms": ["google meet", "zoom", "online class", "teams"],
    "Presentation software usage": ["powerpoint", "slides", "presentation"],
    "Professional development activities": ["training", "workshop", "professional development"],
    "Continuous learning practices": ["learn", "update", "skill"],
    "Technology integration challenges": ["difficult", "challenge", "problem", "barrier"],
    "Resource limitations": ["lack", "limited", "insufficient", "shortage"],
    "Digital assessment practices": ["assessment", "evaluation", "quiz", "test"],
    "Feedback and grading": ["feedback", "grade", "marking"],
    "Student engagement strategies": ["engage", "motivate", "interest"],
    "Interactive teaching methods": ["interaction", "participation", "active"],
    "Technology-related practice": ["technology", "digital", "computer"],
    "Teaching-related practice": ["student", "class", "teach"]
}


def load_codebook(path: Optional[Path] = None) -> Dict[str, List[str]]:
    """
    Load codebook from JSON file. If none provided, return DEFAULT_CODEBOOK.
    """
    if path:
        try:
            with open(path, 'r', encoding='utf-8') as f:
                cb = json.load(f)
            logger.info(f"Loaded codebook from {path}")
            return cb
        except Exception as e:
            logger.warning(f"Failed to load codebook at {path}: {e}. Using default.")
    return DEFAULT_CODEBOOK


def extract_text_from_file(file_path: Path) -> str:
    """
    Extract text from .docx, .pdf, .txt.
    Returns extracted text (empty string if none).
    """
    try:
        suffix = file_path.suffix.lower()
        if suffix == '.docx':
            doc = docx.Document(file_path)
            return '\n'.join(p.text for p in doc.paragraphs if p.text)
        elif suffix == '.pdf':
            text = []
            with open(file_path, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                for page in reader.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text.append(page_text)
            return '\n'.join(text)
        elif suffix == '.txt':
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()
        else:
            logger.warning(f"Unsupported format: {file_path}")
            return ""
    except Exception as e:
        logger.error(f"Error extracting {file_path}: {e}")
        return ""


def extract_participant_responses(
    transcript_text: str,
    speaker_markers: Optional[List[str]] = None,
    interviewer_markers: Optional[List[str]] = None,
    min_length: int = 20
) -> List[str]:
    """
    Extract participant (respondent) segments from a transcript string.
    """
    if speaker_markers is None:
        speaker_markers = ['participant:', 'interviewee:', 'teacher:', 'respondent:']
    if interviewer_markers is None:
        interviewer_markers = ['researcher:', 'interviewer:', 'moderator:']

    lines = [ln.strip() for ln in transcript_text.splitlines() if ln.strip()]
    participant_responses = []
    current = []
    is_participant = None  # unknown initially

    for line in lines:
        low = line.lower()
        # detect explicit markers
        if any(low.startswith(m) for m in speaker_markers):
            if current and is_participant:
                participant_responses.append(' '.join(current).strip())
            is_participant = True
            # remove marker prefix (like "Participant:")
            cleaned = re.sub(r'^\w+:\s*', '', line)
            current = [cleaned] if cleaned else []
        elif any(low.startswith(m) for m in interviewer_markers):
            if current and is_participant:
                participant_responses.append(' '.join(current).strip())
            is_participant = False
            current = []
        else:
            # no explicit marker: use content-based heuristics
            if is_participant is None:
                # More robust heuristics based on line content rather than alternation
                has_personal_pronouns = bool(re.search(r'\b(I|we|my|our|us|me)\b', line, re.I))
                has_question_words = bool(re.search(r'\b(what|how|why|when|where|who|could you|would you|can you)\b', line, re.I))
                is_long_statement = len(line) > 60
                
                # Participant responses typically have personal pronouns and are longer statements
                # Interviewer questions typically have question words and are shorter
                if has_personal_pronouns and not has_question_words:
                    is_participant = True
                elif has_question_words or len(line) < 20:
                    is_participant = False
                elif is_long_statement:
                    is_participant = True
                else:
                    # Default to False (safer for data quality - excludes ambiguous lines)
                    is_participant = False
            if is_participant:
                current.append(line)
            else:
                if current:
                    participant_responses.append(' '.join(current).strip())
                current = []

    if current and is_participant:
        participant_responses.append(' '.join(current).strip())

    participant_responses = [r for r in participant_responses if len(r) >= min_length]
    logger.info(f"Extracted {len(participant_responses)} participant response segments")
    return participant_responses


def split_into_sentences(text: str) -> List[str]:
    """
    Simple sentence splitter: splits on ., ?, ! followed by space + capital letter.
    Fallback: split by newline.
    """
    sentences = re.split(r'(?<=[.!?])\s+(?=[A-Z0-9])', text.strip())
    if len(sentences) <= 1:
        sentences = [s.strip() for s in text.splitlines() if s.strip()]
    return [s.strip() for s in sentences if s.strip()]


def suggest_keywords_from_texts(texts: List[str], top_n: int = 20, ngram_range=(1, 2)) -> List[str]:
    """
    Suggest domain keywords using TF-IDF across a list of texts (uploaded transcripts).
    Returns top_n candidate tokens (unigrams and bigrams).
    """
    if not texts:
        return []
    try:
        vectorizer = TfidfVectorizer(stop_words='english', ngram_range=ngram_range, max_df=0.85)
        X = vectorizer.fit_transform(texts)
        scores = X.sum(axis=0).A1  # sum TF-IDF scores across docs
        feature_names = vectorizer.get_feature_names_out()
        ranked_idx = scores.argsort()[::-1]
        suggested = []
        for idx in ranked_idx:
            token = feature_names[idx]
            # basic cleaning: skip tokens with numbers or very short
            if len(token) < 2 or any(ch.isdigit() for ch in token):
                continue
            suggested.append(token)
            if len(suggested) >= top_n:
                break
        logger.info(f"Suggested {len(suggested)} keywords via TF-IDF")
        return suggested
    except Exception as e:
        logger.error(f"Keyword suggestion failed: {e}")
        return []


def generate_initial_code(text: str, codebook: Dict[str, List[str]], domain_keywords: Optional[List[str]] = None) -> Tuple[str, Optional[str]]:
    """
    Return the first matching code label based on keywords in codebook.
    If a domain keyword is matched, return a domain-specific code and the matched keyword in notes.
    Returns: (code_label, matched_domain_keyword_or_None)
    """
    text_lower = text.lower()

    # 1) Domain-specific keywords (priority)
    if domain_keywords:
        for kw in domain_keywords:
            if kw.lower() in text_lower:
                # Return a standardized domain-specific label and the matched kw
                return f"Domain-specific practice ({kw})", kw

    # 2) Regular codebook matching
    for label, keywords in codebook.items():
        for kw in keywords:
            if kw.lower() in text_lower:
                return label, None

    # heuristics fallback
    if any(w in text_lower for w in ['technology', 'digital', 'computer']):
        return "Technology-related practice", None
    if any(w in text_lower for w in ['student', 'class', 'teach', 'learner']):
        return "Teaching-related practice", None
    return "General educational practice", None


def stage1_initial_coding(
    transcript_text: str,
    interview_id: str,
    codebook: Dict[str, List[str]],
    domain_keywords: Optional[List[str]] = None
) -> pd.DataFrame:
    """
    Stage 1: extract participant responses, split into meaning units, assign initial codes.
    Returns DataFrame with columns: Segment_ID, Interview_Text, Initial_Code, Notes
    """
    participant_responses = extract_participant_responses(transcript_text)
    rows = []
    seg_id = 1
    for resp in participant_responses:
        sentences = split_into_sentences(resp)
        for sent in sentences:
            if len(sent) < 15:
                continue
            code, matched_kw = generate_initial_code(sent, codebook, domain_keywords)
            note = f"Matched domain keyword: {matched_kw}" if matched_kw else ""
            rows.append({
                'Segment_ID': f'S{seg_id:03d}',
                'Interview_Text': sent,
                'Initial_Code': code,
                'Notes': note
            })
            seg_id += 1
    df = pd.DataFrame(rows)
    logger.info(f"Stage1: {len(df)} segments coded for {interview_id}")
    return df


def stage2_code_grouping(stage1_df: pd.DataFrame) -> pd.DataFrame:
    """
    Stage 2: group similar initial codes into broader groups by simple keyword mapping.
    Returns DataFrame: Group_ID, Group_Title, Codes_Included, Segment_IDs, Number_of_Codes
    """
    groups = [
        ('G01', 'Professional Background and Identity', ['professional', 'identity', 'experience', 'career']),
        ('G02', 'Digital Tools and Platforms', ['lms', 'multimedia', 'social media', 'virtual', 'presentation', 'technology']),
        ('G03', 'Professional Development and Learning', ['professional development', 'continuous learning', 'training', 'workshop']),
        ('G04', 'Technology Integration Challenges', ['challenge', 'limitation', 'barrier', 'problem', 'resource']),
        ('G05', 'Assessment and Feedback Practices', ['assessment', 'feedback', 'grading', 'evaluation']),
        ('G06', 'Student Engagement and Interaction', ['engagement', 'interaction', 'motivation', 'participation']),
        ('G07', 'General Teaching Practices', ['teaching', 'educational', 'general'])
    ]
    stage2_rows = []
    for gid, title, keywords in groups:
        matched_codes = []
        segment_ids = []
        for _, row in stage1_df.iterrows():
            ic = str(row['Initial_Code'])
            if any(k in ic.lower() for k in keywords):
                if ic not in matched_codes:
                    matched_codes.append(ic)
                segment_ids.append(row['Segment_ID'])
        if matched_codes:
            stage2_rows.append({
                'Group_ID': gid,
                'Group_Title': title,
                'Codes_Included': ', '.join(sorted(set(matched_codes))),
                'Segment_IDs': ', '.join(segment_ids),
                'Number_of_Codes': len(segment_ids)
            })
    df2 = pd.DataFrame(stage2_rows)
    logger.info(f"Stage2: {len(df2)} groups created")
    return df2


def stage3_thematic_framework(stage1_df: pd.DataFrame, research_questions: List[str]) -> pd.DataFrame:
    """
    Stage 3: build thematic framework mapping initial codes to RQs
    Returns DataFrame: Research_Question, Main_Theme, Sub_Theme, Supporting_Code, Supporting_Quote, Segment_ID
    """
    rq_mappings = {
        'Technology Integration Barriers': ['challenge', 'limitation', 'barrier', 'problem'],
        'Skill and Knowledge Gaps': ['lack', 'limited', 'insufficient', 'gap'],
        'Institutional Constraints': ['resource', 'support', 'institutional'],
        'Professional Development Strategies': ['training', 'workshop', 'development', 'learning'],
        'Peer and Collaborative Learning': ['colleague', 'peer', 'collaboration'],
        'Self-Directed Learning': ['self', 'independent', 'personal'],
        'Communication and Collaboration Tools': ['lms', 'social media', 'virtual', 'communication'],
        'Multimedia and Presentation Tools': ['multimedia', 'presentation', 'video', 'visual'],
        'Assessment and Feedback Systems': ['assessment', 'feedback', 'evaluation', 'grading'],
        'Student Engagement Technologies': ['engagement', 'interaction', 'motivation']
    }

    rows = []
    for theme_name, keywords in rq_mappings.items():
        relevant = []
        for _, r in stage1_df.iterrows():
            if any(k in str(r['Initial_Code']).lower() for k in keywords):
                relevant.append({'code': r['Initial_Code'], 'quote': r['Interview_Text'], 'segment_id': r['Segment_ID']})
        if not relevant:
            continue
        if research_questions:
            rq_idx = (abs(hash(theme_name)) % len(research_questions))
            rq_text = research_questions[rq_idx]
        else:
            rq_text = f"(No RQ) — {theme_name}"
        for i, item in enumerate(relevant[:5]):
            rows.append({
                'Research_Question': rq_text,
                'Main_Theme': theme_name,
                'Sub_Theme': f"{theme_name} - Example {i+1}",
                'Supporting_Code': item['code'],
                'Supporting_Quote': item['quote'][:500],
                'Segment_ID': item['segment_id']
            })
    df3 = pd.DataFrame(rows)
    logger.info(f"Stage3: {len(df3)} thematic entries created")
    return df3


def create_excel_file(df: pd.DataFrame, file_path: Path, sheet_name: str = 'Sheet1'):
    """
    Write DataFrame to Excel and do minimal header styling.
    """
    file_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)
    wb.save(file_path)
    logger.info(f"Excel saved: {file_path}")


def process_single_transcript(
    file_path: Path,
    output_folder: Path,
    codebook: Dict[str, List[str]],
    research_questions: List[str],
    domain_keywords: Optional[List[str]] = None
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Process a single transcript file through Stage1-3 and write excel files to disk.
    domain_keywords: optional list of domain-specific keywords to prioritize.
    """
    interview_id = file_path.stem
    text = extract_text_from_file(file_path)
    if not text:
        logger.warning(f"No text for {file_path}")
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    stage1 = stage1_initial_coding(text, interview_id, codebook, domain_keywords)
    stage2 = stage2_code_grouping(stage1) if not stage1.empty else pd.DataFrame()
    stage3 = stage3_thematic_framework(stage1, research_questions) if not stage1.empty else pd.DataFrame()

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_base = output_folder / f"{interview_id}_{timestamp}"
    out_base.mkdir(parents=True, exist_ok=True)

    if not stage1.empty:
        create_excel_file(stage1, out_base / f"{interview_id}_Stage1_Initial_Coding.xlsx", sheet_name="Initial Coding")
    if not stage2.empty:
        create_excel_file(stage2, out_base / f"{interview_id}_Stage2_Code_Grouping.xlsx", sheet_name="Code Grouping")
    if not stage3.empty:
        create_excel_file(stage3, out_base / f"{interview_id}_Stage3_Thematic_Framework.xlsx", sheet_name="Thematic Framework")

    return stage1, stage2, stage3


def make_output_folder(project_name: str) -> Path:
    """
    Create an outputs folder named by project and timestamp.
    """
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    p = Path('outputs') / f"{project_name}_{ts}"
    p.mkdir(parents=True, exist_ok=True)
    return p
